import streamlit as st
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from scipy.stats import studentized_range
from matplotlib import font_manager
import io
import itertools
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 共通設定・関数
# ==========================================
st.set_page_config(page_title="栽培試験データ統合プラットフォーム", page_icon="🌱", layout="wide")

def set_japanese_font():
    try:
        import japanize_matplotlib
        japanize_matplotlib.japanize()
        return
    except ImportError:
        pass
    candidates = ['IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP',
                  'Hiragino Sans', 'Hiragino Maru Gothic Pro', 'MS Gothic',
                  'Yu Gothic', 'Meiryo']
    available = {f.name for f in font_manager.fontManager.ttflist}
    for font in candidates:
        if font in available:
            plt.rcParams['font.family'] = font
            return

set_japanese_font()

# ==========================================
# 修正1: CLDアルゴリズム（Piepho sweep法）
# ==========================================
def get_cld_letters(groups_sorted, tukey_df):
    """
    Piepho (2004) sweep法に基づく正しいCLD実装。
    groups_sorted: 平均値降順のグループ名リスト（str）
    tukey_df: group1, group2, reject 列を持つDataFrame
    """
    # 有意差なしペアをセットに格納
    ns_pairs = set()
    for _, row in tukey_df.iterrows():
        g1, g2 = str(row['group1']), str(row['group2'])
        if not row['reject']:
            ns_pairs.add((g1, g2))
            ns_pairs.add((g2, g1))
    for g in groups_sorted:
        ns_pairs.add((g, g))  # 自分自身

    letters = {g: [] for g in groups_sorted}
    letter_groups = []   # 既に割り当て済みの「グループセット」
    current_idx = 0

    for gi in groups_sorted:
        # gi と有意差のない全グループを「吸収セット」として列挙
        absorb = [gj for gj in groups_sorted if (gi, gj) in ns_pairs]

        # 同じ吸収セットが既に登録済みかチェック
        matched = any(set(lg) == set(absorb) for lg in letter_groups)

        if not matched:
            new_letter = chr(ord('a') + current_idx)
            current_idx += 1
            letter_groups.append(absorb)
            for g in absorb:
                letters[g].append(new_letter)

    return {g: ''.join(sorted(set(v))) for g, v in letters.items()}


# ==========================================
# 修正7: ANOVAモデルの残差MSを使ったTukey検定
# ==========================================
def tukey_from_anova_model(model, factor, df, target):
    """
    多要因ANOVAモデルの残差MSとdf_residを使ったTukey検定。
    pairwise_tukeyhsd（一元配置前提）より多要因モデルに整合する。
    """
    groups = sorted(df[factor].astype(str).unique())
    n_groups = len(groups)
    mse = model.mse_resid
    df_resid = model.df_resid

    group_stats = df.groupby(factor)[target].agg(['mean', 'count'])

    results = []
    for g1, g2 in itertools.combinations(groups, 2):
        m1 = group_stats.loc[g1, 'mean']
        m2 = group_stats.loc[g2, 'mean']
        n1 = group_stats.loc[g1, 'count']
        n2 = group_stats.loc[g2, 'count']
        # 不均等サイズ対応: Tukey-Kramer法
        se = np.sqrt(mse * (1/n1 + 1/n2) / 2)
        q_stat = abs(m1 - m2) / se
        p_val = float(1 - studentized_range.cdf(q_stat, n_groups, df_resid))
        results.append({
            'group1': g1,
            'group2': g2,
            'meandiff': m1 - m2,
            'q_stat': q_stat,
            'p-adj': p_val,
            'reject': p_val < 0.05
        })

    return pd.DataFrame(results)


# ==========================================
# 修正8: Excel出力（openpyxl）
# ==========================================
def build_excel_report(report_anova, tukey_results, t_col, f_cols, formula, ss_type):
    wb = openpyxl.Workbook()

    # --- シート1: 分散分析表 ---
    ws1 = wb.active
    ws1.title = "分散分析表"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    sig_fill    = PatternFill("solid", fgColor="FFCCCC")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font   = Font(bold=True)
    center      = Alignment(horizontal="center")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1.append([f"モデル: {formula}"])
    ws1.append([f"SS タイプ: Type {ss_type}"])
    ws1.append([])

    cols = list(report_anova.columns)
    header_row = ["要因"] + cols
    ws1.append(header_row)
    for cell in ws1[ws1.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for idx, row in report_anova.iterrows():
        clean_idx = str(idx).replace('C(Q("', '').replace('"))', '').replace(':', ' × ')
        row_data = [clean_idx] + [row[c] for c in cols]
        ws1.append(row_data)
        for i, cell in enumerate(ws1[ws1.max_row]):
            cell.border = border
            cell.alignment = center
            if i == 0:
                cell.font = bold_font
            if cols[i-1] == '判定' if i > 0 else False:
                if cell.value in ['*', '**']:
                    cell.fill = sig_fill

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    # --- シート2以降: 要因別多重比較 ---
    for factor, data_dict in tukey_results.items():
        ws = wb.create_sheet(title=f"多重比較_{factor}"[:31])
        ws.append([f"要因: {factor}  目的変数: {t_col}"])
        ws.append(["※ 同じ文字を持つ水準間に有意差なし（Tukey法, α=0.05）"])
        ws.append([])

        df_rep = data_dict['report']
        ws.append(list(df_rep.columns))
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for _, row in df_rep.iterrows():
            ws.append(list(row))
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = center

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ==========================================
# サイドバーナビゲーション
# ==========================================
st.sidebar.title("🌱 メニュー")
app_mode = st.sidebar.radio(
    "機能を選択してください",
    ["📝 1. 調査様式作成 (PlotBuilder)", "📊 2. データ解析 (ANOVA)"]
)
st.sidebar.divider()
st.sidebar.info("💡 **使い方**\n\n"
                "1. 「調査様式作成」で整然データのテンプレートを作り、データを入力します。\n"
                "2. 「データ解析」にそのデータを貼り付けることで、すぐに解析が可能です。")

# ==========================================
# モード1: PlotBuilder (様式作成)
# ==========================================
if app_mode == "📝 1. 調査様式作成 (PlotBuilder)":
    st.title("📝 PlotBuilder - 調査データ様式ジェネレーター")
    st.markdown("試験の要因数と水準を設定すると、統計解析にそのまま使える「整然データ（Tidy Data）」形式の入力様式を自動生成します。")

    with st.expander("💡 新人さんへ：データ入力 4つの鉄則（必ず読んでね！）", expanded=True):
        st.markdown("""
        統計解析をスムーズに行うため、以下のルールを守ってください。

        1. **1行1データの原則**: 「横」に伸ばさず「縦」に追加する。
        2. **セル結合は絶対禁止**: 結合するとプログラムで読み込めなくなります。
        3. **「数値」と「文字」を混ぜない**: 例）「45(病害)」はNG。特記事項は「備考」列へ。
        4. **欠損値の扱いを統一**: 枯死等でデータが取れない場合は「空欄」にする（「-」や「欠」は入力しない）。
        """)

    st.divider()
    st.subheader("1. 試験区の設計")
    num_factors = st.number_input("設定する要因の数", min_value=1, max_value=5, value=2, step=1)

    factor_names = []
    factor_levels_list = []

    st.markdown("#### 要因と水準の入力")
    for i in range(num_factors):
        col1, col2 = st.columns([1, 2])
        with col1:
            default_name = ["品種", "施肥量", "栽植密度"][i] if i < 3 else f"要因{i+1}"
            f_name = st.text_input(f"要因 {i+1} の名前", value=default_name, key=f"name_{i}")
            factor_names.append(f_name)
        with col2:
            default_levels = ["シマアカリ, ニシユタカ, アイユタカ", "少肥, 標準, 多肥", "疎, 密"][i] if i < 3 else "水準A, 水準B"
            f_levels = st.text_input(f"要因 {i+1} の水準（カンマ区切り）", value=default_levels, key=f"levels_{i}")
            cleaned_levels = [x.strip() for x in f_levels.split(",") if x.strip()]
            factor_levels_list.append(cleaned_levels)

    st.markdown("#### 反復と調査項目の設定")
    col3, col4 = st.columns(2)
    with col3:
        reps = st.number_input("反復（ブロック）数", min_value=1, max_value=20, value=3)
    with col4:
        target_var = st.text_input("目的変数（調査項目）", "収量_kg")

    st.divider()
    if st.button("📝 この設計で入力様式を作成する", type="primary"):
        rep_list = list(range(1, reps + 1))
        combinations = list(itertools.product(rep_list, *factor_levels_list))
        columns = ["反復"] + factor_names
        df_template = pd.DataFrame(combinations, columns=columns)
        df_template = df_template.sort_values(by=["反復"] + factor_names).reset_index(drop=True)
        df_template[target_var] = None
        df_template["備考"] = None

        st.subheader("2. 生成された入力様式")
        total_plots = reps
        for levels in factor_levels_list:
            total_plots *= len(levels)

        st.success(f"全 {total_plots} 区画の入力様式が作成されました！")
        st.info("💡 ダウンロードしたCSVに調査データを入力したら、左のメニューから**「データ解析 (ANOVA)」**を開き、データを貼り付けるだけですぐに解析できます。\n\n"
                "⚠️ **「備考」列はANOVA解析の要因に選ばないでください。**")
        st.dataframe(df_template, use_container_width=True)

        csv = df_template.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 この様式をダウンロード (CSV)",
            data=csv,
            file_name="field_survey_template.csv",
            mime="text/csv",
        )

# ==========================================
# モード2: 多要因分散分析＆成績書作成
# ==========================================
elif app_mode == "📊 2. データ解析 (ANOVA)":
    st.title("📊 栽培試験データ 多要因分散分析＆成績書作成ツール")

    # 修正3: session_stateを初期化
    if 'analyzed' not in st.session_state:
        st.session_state.analyzed = False

    data_source = st.radio(
        "データの入力方法：",
        ["📋 Excelからデータを貼り付ける（推奨）", "🥔 サンプルデータ（明確な交互作用あり）で試す"],
        horizontal=True
    )

    df_real = None

    if data_source == "🥔 サンプルデータ（明確な交互作用あり）で試す":
        varieties   = ['シマアカリ', 'ニシユタカ', 'デジマ', 'アイユタカ']
        locations   = ['西之表', '中種子', '南種子', '熊毛']
        fertilizers = ['無施肥', '少肥', '標準', '多肥']
        data = []
        for v, l, f in itertools.product(varieties, locations, fertilizers):
            for _ in range(3):
                base = 1000
                if v == 'シマアカリ': base += 200
                elif v == 'ニシユタカ': base += 100
                if l == '西之表': base += 100
                elif l == '熊毛': base += 80
                if f == '少肥': base += 50
                elif f == '標準': base += 150
                elif f == '多肥': base += 200
                if v == 'シマアカリ' and f == '多肥': base += 300
                data.append([v, l, f, np.random.normal(base, 100)])
        df_real = pd.DataFrame(data, columns=['品種', '地域', '施肥量', '収量'])
        st.success(f"✅ サンプルデータを読み込みました（全{len(df_real)}件）。")

    else:
        st.info("💡 **データの貼り付け手順:**\n\n"
                "1. Excelから解析したいデータ（1行目は列名）をコピーします。\n"
                "2. 下の黒枠内をクリックして **貼り付け（Ctrl + V）** します。\n"
                "3. **`Ctrl` + `Enter`** か **「📥 データを確定して読み込む」** ボタンを押してください。")

        with st.expander("👀 正しいデータ形式（整然データ）の例を見る", expanded=False):
            st.markdown("""
            **【OKな例】要因ごとに列が分かれており、縦にデータが並んでいる状態**
            | 反復 | 品種 | 施肥量 | 収量 |
            | :--- | :--- | :--- | :--- |
            | 1 | ニシユタカ | 少肥 | 1050 |
            | 1 | シマアカリ | 多肥 | 1300 |
            | 2 | ニシユタカ | 少肥 | 1020 |
            """)

        pasted_data = st.text_area("ここにExcelデータを貼り付けてください (貼り付け後、Ctrl+Enter で確定)", height=150)
        if st.button("📥 データを確定して読み込む"):
            pass
        if pasted_data:
            try:
                df_real = pd.read_csv(io.StringIO(pasted_data), sep=None, engine='python')
                st.success("✅ データを正常に読み込みました。下の「解析設定」に進んでください。")
            except Exception as e:
                st.error(f"データの読み込みに失敗しました。コピーした範囲が正しいか確認してください。エラー詳細: {e}")

    # ==========================================
    # 解析設定と実行
    # ==========================================
    if df_real is not None:
        with st.expander("🔍 読み込んだデータのプレビュー", expanded=False):
            st.dataframe(df_real.head(10))

        st.subheader("⚙️ モデルの設定と解析")
        cols = df_real.columns.tolist()

        col1, col2 = st.columns(2)
        with col1:
            target_col = st.selectbox("目的変数（数値データ）", cols, index=len(cols)-1)
        with col2:
            # 修正4: available_factorsで options と default を一致させる
            # 修正10: 「備考」列を候補から除外
            available_factors = [c for c in cols if c != target_col and c != "備考"]
            factor_cols = st.multiselect(
                "主効果とする要因（複数選択可）",
                available_factors,
                default=available_factors
            )

        # 修正6: Type II / Type III の選択
        ss_type = st.radio(
            "平方和のタイプ（SS Type）",
            [2, 3],
            format_func=lambda x: f"Type {x}  " + ("（交互作用なし・推奨デフォルト）" if x == 2 else "（交互作用あり・不均等セルサイズ）"),
            horizontal=True
        )

        possible_interactions = []
        selected_interactions = []
        if len(factor_cols) >= 2:
            possible_interactions = list(itertools.combinations(factor_cols, 2))
            interaction_labels = [f"{c[0]} × {c[1]}" for c in possible_interactions]
            selected_interaction_labels = st.multiselect(
                "考慮する交互作用（オプション・必要なものだけ選択）",
                interaction_labels,
                default=[]
            )
            selected_interactions = [possible_interactions[interaction_labels.index(l)] for l in selected_interaction_labels]

        # 修正5: 反復列選択の警告
        rep_candidates = [c for c in factor_cols if '反復' in c or 'rep' in c.lower() or 'block' in c.lower() or 'Rep' in c]
        if rep_candidates:
            st.warning(f"⚠️ **{', '.join(rep_candidates)}** が要因に含まれています。"
                       "完全無作為化デザイン（CRD）なら除外可。"
                       "乱塊法（RCBD）の場合はブロック効果として含めることは正しいですが、"
                       "Tukey多重比較の対象から外すことを推奨します。")

        # 修正3: 設定変更時にsession_stateをリセット
        current_config = (target_col, tuple(sorted(factor_cols)), tuple(sorted(str(x) for x in selected_interactions)), ss_type)
        if 'last_config' not in st.session_state:
            st.session_state.last_config = None
        if st.session_state.last_config != current_config:
            st.session_state.analyzed = False

        if st.button("🚀 解析を実行する", type="primary") and factor_cols:
            df_clean = df_real.copy()
            df_clean[target_col] = pd.to_numeric(df_clean[target_col], errors='coerce')
            df_clean = df_clean.dropna(subset=[target_col] + factor_cols)
            for f in factor_cols:
                df_clean[f] = df_clean[f].astype(str)

            st.session_state.analyzed = True
            st.session_state.last_config = current_config
            st.session_state.df_clean = df_clean
            st.session_state.target_col = target_col
            st.session_state.factor_cols = factor_cols
            st.session_state.selected_interactions = selected_interactions
            st.session_state.ss_type = ss_type

        if st.session_state.analyzed:
            st.divider()

            df_eval  = st.session_state.df_clean
            t_col    = st.session_state.target_col
            f_cols   = st.session_state.factor_cols
            s_ints   = st.session_state.selected_interactions
            ss_t     = st.session_state.ss_type

            sns.set_theme(style="whitegrid")
            set_japanese_font()

            # --- ANOVA計算 ---
            formula_terms = [f'C(Q("{f}"))' for f in f_cols]
            for f1, f2 in s_ints:
                formula_terms.append(f'C(Q("{f1}")):C(Q("{f2}"))')
            formula = f'Q("{t_col}") ~ ' + ' + '.join(formula_terms)

            anova_success = False
            try:
                model = ols(formula, data=df_eval).fit()
                anova_res = anova_lm(model, typ=ss_t)
                anova_res['mean_sq'] = anova_res['sum_sq'] / anova_res['df']
                anova_res['寄与率(%)'] = (anova_res['sum_sq'] / anova_res['sum_sq'].sum()) * 100

                def get_sig(p):
                    if pd.isna(p): return ""
                    if p < 0.01: return "**"
                    if p < 0.05: return "*"
                    return "ns"

                anova_res['判定'] = anova_res['PR(>F)'].apply(get_sig)
                report_anova = anova_res[['df', 'sum_sq', 'mean_sq', 'F', 'PR(>F)', '判定', '寄与率(%)']].copy()
                report_anova.columns = ['自由度', '平方和(SS)', '平均平方(MS)', 'F値', 'p値', '判定', '寄与率(%)']
                anova_success = True
            except Exception as e:
                st.error(f"分散分析の計算中にエラーが発生しました: {e}")

            # --- 修正7: モデルベースTukey計算 ---
            tukey_results = {}
            # 修正5: 反復列はTukeyの対象外とする
            tukey_targets = [f for f in f_cols if f not in rep_candidates]
            for factor in tukey_targets:
                summary_stats = df_eval.groupby(factor)[t_col].agg(['count', 'mean', 'std']).reset_index()
                summary_stats.rename(columns={'count': 'N', 'mean': '平均値', 'std': '標準偏差'}, inplace=True)
                try:
                    if anova_success:
                        tukey_df = tukey_from_anova_model(model, factor, df_eval, t_col)
                    else:
                        # フォールバック: 一元配置Tukey
                        tukey_obj = pairwise_tukeyhsd(endog=df_eval[t_col], groups=df_eval[factor], alpha=0.05)
                        tukey_df = pd.DataFrame(data=tukey_obj._results_table.data[1:], columns=tukey_obj._results_table.data[0])

                    means_sorted = df_eval.groupby(factor)[t_col].mean().sort_values(ascending=False)
                    groups_sorted = means_sorted.index.astype(str).tolist()
                    cld_map = get_cld_letters(groups_sorted, tukey_df)
                    letters_df = pd.DataFrame({'_g': list(cld_map.keys()), '有意差': list(cld_map.values())})
                    letters_df.rename(columns={'_g': factor}, inplace=True)

                    final_report = pd.merge(summary_stats, letters_df, on=factor)
                    final_report = final_report.sort_values('平均値', ascending=False).reset_index(drop=True)
                    tukey_results[factor] = {'report': final_report, 'order': groups_sorted}
                except Exception as ex:
                    st.warning(f"要因「{factor}」の多重比較でエラー: {ex}")

            # --- タブ表示 ---
            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                "📋 1. 分散分析表",
                "📊 2. 主効果と多重比較",
                "✖️ 3. 交互作用の解析",
                "📥 4. Excel出力",
                "🤖 5. AI解説用プロンプト"
            ])

            # ---- Tab1: 分散分析表 ----
            with tab1:
                if anova_success:
                    st.header("1. 分散分析表 (成績書用フォーマット)")
                    st.caption(f"モデル: `{formula}`  ／  SS Type: {ss_t}")
                    st.dataframe(
                        report_anova.style.format({
                            '自由度': '{:.0f}', '平方和(SS)': '{:.1f}', '平均平方(MS)': '{:.1f}',
                            'F値': '{:.3f}', 'p値': '{:.4f}', '寄与率(%)': '{:.2f}'
                        }).apply(lambda x: ['background-color: #ffcccc' if v in ['*', '**'] else '' for v in x], subset=['判定']),
                        use_container_width=True
                    )

            # ---- Tab2: 主効果と多重比較 ----
            with tab2:
                st.header("2. 主効果の集計表 (Tukeyの多重比較付き)")
                st.caption("※ 多要因ANOVAモデルの残差MSを使用（Tukey-Kramer法）")
                if not tukey_results:
                    st.info("多重比較結果がありません。")
                for factor, data_dict in tukey_results.items():
                    st.subheader(f"▶ 要因: {factor}")
                    col_t1, col_t2 = st.columns([2, 3])

                    with col_t1:
                        st.dataframe(
                            data_dict['report'].style.format({'平均値': '{:.1f}', '標準偏差': '{:.1f}'}),
                            use_container_width=True
                        )

                    with col_t2:
                        fig_box, ax_box = plt.subplots(figsize=(8, 4))
                        sns.boxplot(x=factor, y=t_col, data=df_eval, order=data_dict['order'],
                                    ax=ax_box, color='#f0f0f0', showfliers=False)
                        sns.stripplot(x=factor, y=t_col, data=df_eval, order=data_dict['order'],
                                      ax=ax_box, color='black', alpha=0.5)
                        for gname in data_dict['order']:
                            letter = data_dict['report'].loc[data_dict['report'][factor] == gname, '有意差'].values[0]
                            xpos   = data_dict['order'].index(gname)
                            ymax   = df_eval[df_eval[factor] == gname][t_col].max()
                            ax_box.text(xpos, ymax + abs(ymax) * 0.02, letter,
                                        ha='center', va='bottom', fontweight='bold',
                                        color='#d62728', fontsize=12)
                        ax_box.set_ylabel(t_col)
                        st.pyplot(fig_box)
                        plt.close(fig_box)

            # ---- Tab3: 交互作用図 ----
            with tab3:
                if s_ints:
                    st.header("3. 交互作用図（要因間の相乗効果の確認）")
                    if anova_success:
                        # 修正9: 有意な交互作用のみ強調
                        sig_ints = []
                        for f1, f2 in s_ints:
                            label = f"{f1} × {f2}"
                            for idx in report_anova.index:
                                clean = str(idx).replace('C(Q("', '').replace('"))', '').replace(':', ' × ')
                                if clean == label and report_anova.loc[idx, '判定'] in ['*', '**']:
                                    sig_ints.append(label)
                        if sig_ints:
                            st.success(f"✅ 有意な交互作用: **{', '.join(sig_ints)}**")
                        else:
                            st.warning("⚠️ 選択した交互作用はいずれも有意ではありません（分散分析表参照）。グラフはあくまで参考です。")

                    int_options = [f"{c[0]} × {c[1]}" for c in s_ints]
                    selected_int_plot = st.selectbox("グラフ化する交互作用を選択してください", int_options)
                    plot_x, plot_hue = s_ints[int_options.index(selected_int_plot)]
                    if st.toggle("X軸と色分け(凡例)を入れ替える"):
                        plot_x, plot_hue = plot_hue, plot_x

                    fig_int, ax_int = plt.subplots(figsize=(10, 6))
                    sns.pointplot(x=plot_x, y=t_col, hue=plot_hue, data=df_eval, ax=ax_int,
                                  dodge=True, capsize=.1,
                                  markers=['o', 's', '^', 'D', 'v', '<', '>'],
                                  err_kws={'linewidth': 1})
                    ax_int.set_title(f"【交互作用図】 {plot_x} × {plot_hue}", fontsize=16)
                    ax_int.set_xlabel(plot_x, fontsize=12)
                    ax_int.set_ylabel(f"平均 {t_col}", fontsize=12)
                    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title=plot_hue)
                    plt.tight_layout()
                    st.pyplot(fig_int)
                    plt.close(fig_int)
                else:
                    st.info("💡 解析設定で「交互作用」が選択されていないため、グラフは表示されません。")

            # ---- Tab4: Excel出力（修正8） ----
            with tab4:
                st.header("📥 4. 解析結果をExcelでダウンロード")
                if anova_success and tukey_results:
                    excel_bytes = build_excel_report(report_anova, tukey_results, t_col, f_cols, formula, ss_t)
                    st.download_button(
                        label="📥 解析結果をExcelでダウンロード",
                        data=excel_bytes,
                        file_name="anova_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.caption("シート構成: 「分散分析表」＋要因ごとの「多重比較結果」")
                else:
                    st.warning("解析が完了していないため、Excelを生成できません。")

            # ---- Tab5: AIプロンプト ----
            with tab5:
                st.header("🤖 5. AI解説用プロンプト作成")
                st.markdown("以下をコピーしてClaude / ChatGPT / Gemini 等に貼り付けてください。")
                if anova_success:
                    lines = [
                        "あなたは優秀な農業データアナリストです。",
                        "以下の栽培試験の統計解析結果に基づいて、結果の概要と実践的な考察を作成してください。",
                        "",
                        "### 【解析の前提】",
                        f"- 目的変数: {t_col}",
                        f"- 要因: {', '.join(f_cols)}",
                        f"- SS Type: {ss_t}",
                    ]
                    if s_ints:
                        lines.append(f"- 交互作用: {', '.join([f'{a}×{b}' for a, b in s_ints])}")
                    lines += ["", "### 【分散分析結果】",
                              "※ **(1%有意), *(5%有意), ns(有意差なし)"]
                    for idx, row in report_anova.iterrows():
                        if idx == "Residual": continue
                        clean = str(idx).replace('C(Q("', '').replace('"))', '').replace(':', ' × ')
                        lines.append(f"- {clean}: p={row['p値']:.4f} [{row['判定'] or 'ns'}], 寄与率={row['寄与率(%)']:.1f}%")
                    lines += ["", "### 【多重比較結果（Tukey法）】",
                              "※ 同じ文字の水準間に有意差なし"]
                    for factor, dd in tukey_results.items():
                        lines.append(f"\n▼ {factor}（平均値降順）:")
                        for _, r in dd['report'].iterrows():
                            lines.append(f"  - {r[factor]}: {r['平均値']:.1f}  [{r['有意差']}]")
                    lines += [
                        "",
                        "### 【出力フォーマット】",
                        "1. **結論の要約**: 最も寄与率が高い要因とその解釈。",
                        "2. **水準間の比較**: 有意差アルファベットを使った具体的な記述。",
                    ]
                    if s_ints:
                        lines.append("3. **交互作用の解釈**: 有意な交互作用があれば、特定組み合わせの効果を解説。")
                    lines.append("4. **現場フィードバック**: 次試験や実栽培への具体的なアドバイス。")

                    st.code("\n".join(lines), language="markdown")
                else:
                    st.warning("分散分析が完了していないため、プロンプトを生成できません。")
